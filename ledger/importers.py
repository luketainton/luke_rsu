"""Importers for broker exports.

The raw source key is intentionally based on the export values rather than a row
number: E*TRADE downloads are complete histories and their row ordering can vary.
"""

from __future__ import annotations

import csv
import hashlib
import io
import json
from collections import Counter
from datetime import UTC, date, datetime
from decimal import Decimal, InvalidOperation

from openpyxl import load_workbook

from .models import Broker, Grant, Sale, Vest

ETRADE_BROKER_NAME = "Morgan Stanley E*TRADE"
IMPORT_TYPE_CHOICES = (
    ("etrade_benefit_history", "E*TRADE Benefit History"),
    ("legacy_stock_transactions", "LEGACY Cisco ECC - My Stock Transactions"),
    ("legacy_restricted_stock", "LEGACY Cisco ECC - My Restricted Stock"),
    ("schwab_equity_details", "Schwab - Equity Details"),
    ("schwab_transaction_history", "Schwab - Transaction History"),
)


class UnsupportedImport(ValueError):
    """Raised when an uploaded file is not a recognised broker export."""


def _value(row, field):
    value = row.get(field)
    return value.strip() if isinstance(value, str) else value


def _decimal(value):
    if value in (None, ""):
        return None
    try:
        text = str(value).strip().replace(",", "").replace("$", "")
        if text.startswith("(") and text.endswith(")"):
            text = f"-{text[1:-1]}"
        return Decimal(text)
    except InvalidOperation as exc:
        raise UnsupportedImport("The export contains an invalid quantity.") from exc


def _date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if not value:
        return None
    for pattern in ("%m/%d/%Y", "%m-%d-%Y", "%Y-%m-%d", "%Y-%b-%d"):
        try:
            return (
                date.fromisoformat(value)
                if pattern == "%Y-%m-%d"
                else datetime.strptime(value, pattern).replace(tzinfo=UTC).date()
            )
        except ValueError:
            continue
    raise UnsupportedImport("The export contains an unrecognised date.")


def _source_key(event_type, grant_number, event_date, units):
    values = [
        "etrade-benefit-history-v1",
        event_type,
        str(grant_number),
        event_date.isoformat(),
        str(units),
    ]
    return hashlib.sha256(json.dumps(values, separators=(",", ":")).encode()).hexdigest()


def _legacy_source_key(import_type, record_type, record_id, event_date, units):
    values = [import_type, record_type, str(record_id), event_date.isoformat(), str(units)]
    return hashlib.sha256(json.dumps(values, separators=(",", ":")).encode()).hexdigest()


def _event_rows(upload):
    workbook = load_workbook(upload, read_only=True, data_only=True)
    worksheet = workbook.active
    rows = worksheet.iter_rows(values_only=True)
    headers = next(rows, None)
    if not headers:
        raise UnsupportedImport("The workbook is empty.")
    names = [str(header).strip() if header is not None else "" for header in headers]
    required = {"Record Type", "Event Type", "Grant Number"}
    if not required.issubset(names):
        raise UnsupportedImport("This is not an E*TRADE Benefit History export.")
    return [dict(zip(names, row, strict=False)) for row in rows]


def _legacy_rows(upload, required, message):
    upload.seek(0)
    workbook = load_workbook(upload, read_only=True, data_only=True)
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    for index, header_row in enumerate(rows):
        names = [str(header).strip() if header is not None else "" for header in header_row]
        if required.issubset(names):
            return [dict(zip(names, row, strict=False)) for row in rows[index + 1 :]]
    raise UnsupportedImport(message)


def _schwab_transaction_rows(upload):
    upload.seek(0)
    try:
        content = upload.read().decode("utf-8-sig")
    except (AttributeError, UnicodeDecodeError) as exc:
        raise UnsupportedImport("This is not a Schwab Transaction History CSV export.") from exc
    reader = csv.DictReader(io.StringIO(content))
    required = {"Date", "Action", "Quantity"}
    if not reader.fieldnames or not required.issubset(reader.fieldnames):
        raise UnsupportedImport("This is not a Schwab Transaction History CSV export.")

    records = []
    pending = None
    for row in reader:
        if _value(row, "Date") or _value(row, "Action"):
            if pending is not None:
                records.append(pending)
            pending = dict(row)
        elif pending is not None:
            pending.update(
                {field: value for field, value in row.items() if value not in (None, "")}
            )
    if pending is not None:
        records.append(pending)
    return records


def _broker_for_name(workspace, name):
    name = name.strip() if isinstance(name, str) else name
    if not name:
        return None
    broker = Broker.objects.filter(workspace=workspace, name__iexact=name).first()
    return broker or Broker.objects.create(workspace=workspace, name=name)


def _empty_counts():
    return Counter(grants=0, vests=0, sales=0, duplicates=0, skipped=0, missing_sale_prices=0)


def import_legacy_stock_transactions(upload, workspace):
    """Import vest transactions from a legacy Cisco ECC stock-transactions export."""
    rows = _legacy_rows(
        upload,
        {"Date of Transaction", "Transaction Type", "Shares Exercised/Vested", "Grant Number"},
        "This is not a legacy Cisco ECC My Stock Transactions export.",
    )
    counts = _empty_counts()
    unrecognised_types = Counter()
    for row in rows:
        transaction_type = _value(row, "Transaction Type")
        if transaction_type != "Lapse":
            if transaction_type not in (None, "Total"):
                unrecognised_types[str(transaction_type)] += 1
            continue
        event_date = _date(_value(row, "Date of Transaction"))
        units = _decimal(_value(row, "Shares Exercised/Vested"))
        grant_number = _value(row, "Grant Number")
        if not event_date or units is None or units <= 0:
            counts["skipped"] += 1
            continue
        key = _legacy_source_key(
            "legacy-stock-transactions-v1", "lapse", grant_number, event_date, units
        )
        _, created = Vest.objects.get_or_create(
            workspace=workspace,
            source_key=key,
            defaults={
                "date": event_date,
                "units": units,
                "grant_id": str(grant_number or ""),
                "usd_price": _decimal(_value(row, "Sale Price/FMV")),
                "withheld_units": _decimal(_value(row, "Shares used for Taxes")) or Decimal(),
                "income_tax": _decimal(_value(row, "Taxes")) or Decimal(),
                "notes": "Imported from legacy Cisco ECC My Stock Transactions.",
            },
        )
        counts["vests" if created else "duplicates"] += 1
    counts["unrecognised_transaction_types"] = ", ".join(sorted(unrecognised_types))
    return counts


def import_legacy_restricted_stock(upload, workspace):
    """Import award rows from a legacy Cisco ECC restricted-stock export."""
    rows = _legacy_rows(
        upload,
        {"Status", "Award Number", "Award Date", "Shares Awarded", "Selected Broker"},
        "This is not a legacy Cisco ECC My Restricted Stock export.",
    )
    counts = _empty_counts()
    for row in rows:
        if _value(row, "Status") == "Total":
            continue
        award_date = _date(_value(row, "Award Date"))
        units = _decimal(_value(row, "Shares Awarded"))
        award_number = _value(row, "Award Number")
        if not award_date or units is None or units <= 0:
            counts["skipped"] += 1
            continue
        key = _legacy_source_key(
            "legacy-restricted-stock-v1", "award", award_number, award_date, units
        )
        _, created = Grant.objects.get_or_create(
            workspace=workspace,
            source_key=key,
            defaults={
                "date": award_date,
                "units": units,
                "grant_id": str(award_number or ""),
                "usd_price": _decimal(_value(row, "Conversion Price*")),
                "broker": _broker_for_name(workspace, _value(row, "Selected Broker")),
                "notes": "Imported from legacy Cisco ECC My Restricted Stock.",
            },
        )
        counts["grants" if created else "duplicates"] += 1
    return counts


def import_schwab_equity_details(upload, workspace):
    """Import acquired restricted-stock tranches from a Schwab Equity Details export."""
    rows = _legacy_rows(
        upload,
        {"Award Date", "Symbol", "Award ID", "Date Acquired", "Acquisition Price", "Shares"},
        "This is not a Schwab Equity Details export.",
    )
    counts = _empty_counts()
    broker = _broker_for_name(workspace, "Charles Schwab")
    for row in rows:
        if _value(row, "Award ID") in (None, "", "Totals"):
            continue
        event_date = _date(_value(row, "Date Acquired"))
        units = _decimal(_value(row, "Shares"))
        award_id = _value(row, "Award ID")
        available = _decimal(_value(row, "Available to Sell")) or Decimal()
        if not event_date or units is None or units <= 0:
            counts["skipped"] += 1
            continue
        key = _legacy_source_key(
            "schwab-equity-details-v1",
            "acquired",
            f"{award_id}:{_value(row, 'Acquisition Price')}",
            event_date,
            units,
        )
        _, created = Vest.objects.get_or_create(
            workspace=workspace,
            source_key=key,
            defaults={
                "date": event_date,
                "units": units,
                "grant_id": str(award_id),
                "broker": broker,
                "usd_price": _decimal(_value(row, "Acquisition Price")),
                "withheld_units": max(Decimal(), units - available),
                "notes": "Imported from Schwab Equity Details.",
            },
        )
        counts["vests" if created else "duplicates"] += 1
    return counts


def import_schwab_transaction_history(upload, workspace):
    """Import RS deposits and share sales from a Schwab transaction-history CSV."""
    rows = _schwab_transaction_rows(upload)
    counts = _empty_counts()
    unrecognised_actions = Counter()
    broker = _broker_for_name(workspace, "Charles Schwab")
    for row in rows:
        action = _value(row, "Action")
        event_date = _date(_value(row, "Date"))
        units = _decimal(_value(row, "Quantity") or _value(row, "Shares"))
        if action == "Deposit" and (
            _value(row, "Type") == "RS" or _value(row, "Description") == "RS"
        ):
            if not event_date or units is None or units <= 0:
                counts["skipped"] += 1
                continue
            award_id = _value(row, "AwardId") or ""
            key = _legacy_source_key(
                "schwab-transaction-history-v1",
                "deposit",
                f"{award_id}:{_value(row, 'PurchaseDate')}",
                event_date,
                units,
            )
            _, created = Vest.objects.get_or_create(
                workspace=workspace,
                source_key=key,
                defaults={
                    "date": event_date,
                    "units": units,
                    "grant_id": award_id,
                    "broker": broker,
                    "usd_price": _decimal(_value(row, "PurchasePrice")),
                    "notes": "Imported from Schwab Transaction History; deposit quantity is shares received.",
                },
            )
            counts["vests" if created else "duplicates"] += 1
        elif action == "Sale":
            if not event_date or units is None or units <= 0:
                counts["skipped"] += 1
                continue
            grant_id = _value(row, "GrantId") or ""
            sale_price = _decimal(_value(row, "SalePrice"))
            key = _legacy_source_key(
                "schwab-transaction-history-v1",
                "sale",
                f"{grant_id}:{sale_price}",
                event_date,
                units,
            )
            _, created = Sale.objects.get_or_create(
                workspace=workspace,
                source_key=key,
                defaults={
                    "date": event_date,
                    "units": units,
                    "grant_id": grant_id,
                    "broker": broker,
                    "usd_price": sale_price,
                    "notes": "Imported from Schwab Transaction History.",
                },
            )
            counts["sales" if created else "duplicates"] += 1
        elif action not in (None, ""):
            unrecognised_actions[str(action)] += 1
    counts["unrecognised_transaction_types"] = ", ".join(sorted(unrecognised_actions))
    return counts


def import_etrade_benefit_history(upload, workspace):
    """Insert recognisable E*TRADE events, skipping events already imported."""
    rows = _event_rows(upload)
    broker = Broker.objects.filter(workspace=workspace, name__iexact=ETRADE_BROKER_NAME).first()
    if broker is None:
        broker = Broker.objects.create(workspace=workspace, name=ETRADE_BROKER_NAME)
    released = {}
    for row in rows:
        if _value(row, "Event Type") == "Shares released":
            event_date = _date(_value(row, "Date"))
            units = _decimal(_value(row, "Qty. or Amount"))
            if event_date and units is not None:
                released[(_value(row, "Grant Number"), event_date)] = units

    counts = Counter()
    missing_sale_prices = 0
    seen_grants = set()
    for row in rows:
        event_type = _value(row, "Event Type")
        event_date = _date(_value(row, "Date"))
        units = _decimal(_value(row, "Qty. or Amount"))
        grant_number = _value(row, "Grant Number")
        if event_type not in {"Shares granted", "Shares vested", "Shares sold"}:
            continue
        if not event_date or units is None or units <= 0:
            counts["skipped"] += 1
            continue
        if event_type == "Shares granted":
            model, kind, defaults = Grant, "grants", {}
            seen_grants.add(grant_number)
        elif event_type == "Shares vested":
            model, kind = Vest, "vests"
            withheld = max(Decimal(), units - released.get((grant_number, event_date), Decimal()))
            defaults = {"withheld_units": withheld}
        else:
            model, kind = Sale, "sales"
            defaults = {
                "notes": "Imported from E*TRADE; add the USD sale price before CGT reporting."
            }
            missing_sale_prices += 1
        key = _source_key(event_type, grant_number, event_date, units)
        _, created = model.objects.get_or_create(
            workspace=workspace,
            source_key=key,
            defaults={
                "date": event_date,
                "units": units,
                "grant_id": str(grant_number or ""),
                "broker": broker,
                "notes": f"E*TRADE grant {grant_number}",
                **defaults,
            },
        )
        counts[kind if created else "duplicates"] += 1

    # A few exports contain only the award header instead of a Shares granted event.
    for row in rows:
        grant_number = _value(row, "Grant Number")
        if _value(row, "Record Type") != "Grant" or grant_number in seen_grants:
            continue
        event_date = _date(_value(row, "Grant Date"))
        units = _decimal(_value(row, "Granted Qty."))
        if not event_date or units is None or units <= 0:
            continue
        key = _source_key("Grant header", grant_number, event_date, units)
        _, created = Grant.objects.get_or_create(
            workspace=workspace,
            source_key=key,
            defaults={
                "date": event_date,
                "units": units,
                "grant_id": str(grant_number or ""),
                "broker": broker,
                "notes": f"E*TRADE grant {grant_number}",
            },
        )
        counts["grants" if created else "duplicates"] += 1
    counts["missing_sale_prices"] = missing_sale_prices
    for key in ("grants", "vests", "sales", "duplicates"):
        counts.setdefault(key, 0)
    return counts
